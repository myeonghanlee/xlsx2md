import streamlit as st
import pandas as pd
import openpyxl
import pdfplumber
import io
import olefile

# --- 1. 엑셀 병합 셀 처리 및 AI 친화적 MD 변환 함수 ---
def process_excel(file_bytes, sheet_name):
    # openpyxl로 로드하여 병합된 셀의 데이터를 각 셀에 채워넣음 (AI가 맥락을 잃지 않도록)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[sheet_name]
    
    # 병합된 셀 범위 확인 및 값 채우기
    mcr_list = list(ws.merged_cells.ranges)
    for mcr in mcr_list:
        min_col, min_row, max_col, max_row = mcr.bounds
        top_left_cell_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(mcr))
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.value = top_left_cell_value

    # 정제된 데이터를 pandas로 읽어 Markdown 표로 변환
    data = ws.values
    cols = next(data)[0:]
    df = pd.DataFrame(data, columns=cols)
    
    # 빈 값 처리 및 깔끔한 MD 포맷팅
    df = df.fillna("")
    return df.to_markdown(index=False)

# --- 2. PDF 변환 함수 ---
def process_pdf(file_bytes):
    md_text = ""
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            # 텍스트 추출
            text = page.extract_text()
            if text:
                md_text += text + "\n\n"
            # 표 추출 (AI가 인식할 수 있는 MD 표 형태로 변환)
            tables = page.extract_tables()
            for table in tables:
                df = pd.DataFrame(table[1:], columns=table[0])
                df = df.fillna("")
                md_text += df.to_markdown(index=False) + "\n\n"
    return md_text

# --- 3. HWP 변환 함수 (기본 텍스트 추출) ---
def process_hwp(file_bytes):
    # Linux(Streamlit Cloud) 환경에서 HWP 추출은 제한적이나, 내부 텍스트 스트림을 뽑아내는 기본 로직입니다.
    f = olefile.OleFileIO(io.BytesIO(file_bytes))
    dirs = f.listdir()
    
    if ["PrvText"] in dirs:
        text = f.openstream("PrvText").read()
        return text.decode("utf-16le", errors="ignore")
    else:
        return "HWP 파일에서 텍스트를 추출할 수 없습니다. (표 구조가 복잡한 HWP는 PDF 변환 후 업로드 권장)\n"

# --- 웹앱 UI 및 메인 로직 ---
st.set_page_config(page_title="AI 분석용 다중 파일 MD 변환기", layout="wide")
st.title("📄 AI 분석용 문서 변환기 (Excel, PDF, HWP ➡️ MD)")

uploaded_files = st.file_uploader("변환할 파일들을 모두 올려주세요.", type=['xlsx', 'pdf', 'hwp'], accept_multiple_files=True)

if uploaded_files:
    # 엑셀 파일이 있는지 확인하고 첫 번째 엑셀 파일의 시트 목록을 가져옴
    excel_files = [f for f in uploaded_files if f.name.endswith('.xlsx')]
    selected_sheet = None
    
    if excel_files:
        st.subheader("📊 엑셀 시트 선택 (모든 엑셀 파일에 동일하게 적용됩니다)")
        # 첫 번째 엑셀 파일의 시트명 추출
        wb = openpyxl.load_workbook(io.BytesIO(excel_files[0].getvalue()), read_only=True)
        sheet_names = wb.sheetnames
        selected_sheet = st.selectbox("변환할 시트를 선택하세요:", sheet_names)

    if st.button("Markdown으로 변환 시작"):
        for file in uploaded_files:
            file_bytes = file.getvalue()
            filename = file.name
            md_result = f"## 📁 {filename}\n\n"
            
            try:
                if filename.endswith('.xlsx'):
                    # 선택된 시트만 처리
                    md_result += process_excel(file_bytes, selected_sheet)
                elif filename.endswith('.pdf'):
                    md_result += process_pdf(file_bytes)
                elif filename.endswith('.hwp'):
                    md_result += process_hwp(file_bytes)
                
                # 결과 출력 및 다운로드 버튼 제공
                st.markdown(f"### 성공: {filename}")
                with st.expander(f"{filename} 변환 결과 미리보기"):
                    st.text(md_result[:500] + "...\n(미리보기 생략)") # 너무 길면 화면이 꽉 차므로 일부만 표시
                
                st.download_button(
                    label=f"⬇️ {filename}.md 다운로드",
                    data=md_result,
                    file_name=f"{filename}.md",
                    mime="text/markdown"
                )
            except Exception as e:
                st.error(f"{filename} 처리 중 오류 발생: {e}")
